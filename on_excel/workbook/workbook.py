from pathlib import Path

import openpyxl.workbook.workbook as ow
from openpyxl.utils.exceptions import ReadOnlyWorkbookException
from openpyxl.worksheet._write_only import WriteOnlyWorksheet


from ..worksheet.worksheet import Worksheet


class Workbook(ow.Workbook):
    """
    继承了原 Workbook 的类，为了实现扩展方法
    """

    def __init__(self, write_only=False, iso_dates=False,) -> None:
        super(Workbook, self).__init__(write_only, iso_dates)

    def print_self(self):
        print('我是自定义的哒！')

    def find_sheet(self, sheetname):
        sheet_names = self.sheetnames
        find_name = self.rulestring(sheetname)
        for name in sheet_names:
            re_name = self.rulestring(name)
            if((re_name == find_name) | (re_name in find_name) | (find_name in re_name)):
                return name

    def load_sheet(self, sheetname: str, fuzzy: bool = False) -> Worksheet:
        if isinstance(sheetname, str):
            if sheetname in self.sheetnames:
                print('当前表格为：'+sheetname)
                ws = self[sheetname]
                return ws
            elif fuzzy:
                _sheet_name = self.find_sheet(sheetname)
                return self.load_sheet(_sheet_name, False)
        else:
            raise TypeError("sheetname must be not empty str.")

    def create_sheet(self, sheetname):
        if sheetname in self.sheetnames:
            raise ValueError("sheet {0} is exist.".format(sheetname))
        ws = self.create_sheet(sheetname)
        return ws

    def copy_model_sheet(self, modelname, sheetname):
        if modelname in self.sheetnames:
            moudlesheet = self[modelname]
            copyws = self.copy_worksheet(moudlesheet)
            copyws.title = sheetname
            return copyws
        else:
            raise ValueError('没有名为《'+modelname+'》的模板表格')

    def apart_sheets_as_books(self, folder: str, pattern: str = "{0}", col_list=None, callback=None):
        for name in self.sheetnames:
            from_worksheet = self.load_sheet(name)
            df = from_worksheet.unmerge_cell(True).toDataframe()
            out = Workbook()
            to_worksheet = out.create_sheet(title=name)
            # 数据转移
            to_worksheet.append_df(df)
            # 样式转移
            to_worksheet.sys_style(from_worksheet)
            # 自动列宽
            to_worksheet.auto_width(col_list)
            # 额外操作，用函数传入
            if callback:
                to_worksheet = callback(to_worksheet)
            if isinstance(to_worksheet, Worksheet):
                path = Path(folder).joinpath(pattern.format(name))
                if path.exists() and path.is_dir():
                    out.save(str(path))
                else:
                    NotADirectoryError("folder is not a directory.")
            else:
                tip = 'arg function func_call return not Sheet Type!'
                raise TypeError(tip)

    def save(self, path):
        if path == None:
            raise ValueError("path can not be None.")

        print('<'+path+'>工作簿正在保存')
        if 'Sheet' in self.sheetnames:
            self.remove(self['Sheet'])
        super().save(path)
        self.close()
        print('<'+path+'>工作簿保存成功')

    # 重写此方法，用于使用自定义的 Worksheet
    def create_sheet(self, title=None, index=None):
        """Create a worksheet (at an optional index).

        :param title: optional title of the sheet
        :type title: str
        :param index: optional position at which the sheet will be inserted
        :type index: int

        """
        if self.read_only:
            raise ReadOnlyWorkbookException(
                'Cannot create new sheet in a read-only workbook')

        if self.write_only:
            new_ws = WriteOnlyWorksheet(parent=self, title=title)
        else:
            new_ws = Worksheet(parent=self, title=title)

        self._add_sheet(sheet=new_ws, index=index)
        return new_ws
