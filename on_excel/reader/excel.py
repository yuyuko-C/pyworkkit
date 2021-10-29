
# Python stdlib imports
from zipfile import ZipFile, ZIP_DEFLATED, BadZipfile
from sys import exc_info
from io import BytesIO
import os.path
import warnings

from openpyxl.pivot.table import TableDefinition

# # Allow blanket setting of KEEP_VBA for testing
# try:
#     from ..tests import KEEP_VBA
# except ImportError:
#     KEEP_VBA = False


# package imports
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.xml.constants import (
    ARC_SHARED_STRINGS,
    ARC_CORE,
    ARC_CONTENT_TYPES,
    ARC_WORKBOOK,
    ARC_THEME,
    COMMENTS_NS,
    SHARED_STRINGS,
    EXTERNAL_LINK,
    XLTM,
    XLTX,
    XLSM,
    XLSX,
)
from openpyxl.comments.comment_sheet import CommentSheet

from openpyxl.reader.strings import read_string_table
from openpyxl.reader.workbook import WorkbookParser
from openpyxl.styles.stylesheet import apply_stylesheet

from openpyxl.packaging.core import DocumentProperties
from openpyxl.packaging.manifest import Manifest, Override

from openpyxl.packaging.relationship import (
    RelationshipList,
    get_dependents,
    get_rels_path,
)

from openpyxl.worksheet._read_only import ReadOnlyWorksheet
import openpyxl.worksheet._reader as o_reader
from openpyxl.chartsheet import Chartsheet
from openpyxl.worksheet.table import Table
from openpyxl.drawing.spreadsheet_drawing import SpreadsheetDrawing
import openpyxl.reader.excel as o_excel

from openpyxl.xml.constants import (
    SHEET_MAIN_NS,
    EXT_TYPES,
)

from openpyxl.xml.functions import fromstring

from openpyxl.reader.drawings import find_images


SUPPORTED_FORMATS = ('.xlsx', '.xlsm', '.xltx', '.xltm')



import openpyxl.reader.excel as o_excel


from on_excel.worksheet._reader import WorksheetReader
from on_excel.reader.workbook import WorkbookParser
from on_excel.cell import MergedCell



class ExcelReader(o_excel.ExcelReader):
    """
    继承了原 ExcelReader 的类，为了使用自定义的 Workbook
    """

    # 重写此方法，用于使用自定义的 WorkbookParser，从而使用自定义的 Workbook
    def read_workbook(self):
        wb_part = o_excel._find_workbook_part(self.package)
        self.parser = WorkbookParser(
            self.archive, wb_part.PartName[1:], keep_links=self.keep_links)
        self.parser.parse()
        wb = self.parser.wb
        wb._sheets = []
        wb._data_only = self.data_only
        wb._read_only = self.read_only
        wb.template = wb_part.ContentType in (XLTX, XLTM)
        # wb.file_path = self.file_path

        # If are going to preserve the vba then attach a copy of the archive to the
        # workbook so that is available for the save.
        if self.keep_vba:
            wb.vba_archive = ZipFile(BytesIO(), 'a', ZIP_DEFLATED)
            for name in self.valid_files:
                wb.vba_archive.writestr(name, self.archive.read(name))

        if self.read_only:
            wb._archive = self.archive

        self.wb = wb

    # 重写此方法，用于使用自定义的 WorksheetReader,从而使用自定义的 Cell
    def read_worksheets(self):
        comment_warning = """Cell '{0}':{1} is part of a merged range but has a comment which will be removed because merged cells cannot contain any data."""
        for sheet, rel in self.parser.find_sheets():
            if rel.target not in self.valid_files:
                continue

            if "chartsheet" in rel.Type:
                self.read_chartsheet(sheet, rel)
                continue

            rels_path = get_rels_path(rel.target)
            rels = RelationshipList()
            if rels_path in self.valid_files:
                rels = get_dependents(self.archive, rels_path)

            if self.read_only:
                ws = ReadOnlyWorksheet(
                    self.wb, sheet.name, rel.target, self.shared_strings)
                self.wb._sheets.append(ws)
                continue
            else:
                fh = self.archive.open(rel.target)
                # 因 create_sheet 在 Workbook 中重写，所以生成的也是自定义的 Worksheet
                ws = self.wb.create_sheet(sheet.name)
                ws._rels = rels
                ws_parser = WorksheetReader(
                    ws, fh, self.shared_strings, self.data_only)
                ws_parser.bind_all()

            # assign any comments to cells
            for r in rels.find(COMMENTS_NS):
                src = self.archive.read(r.target)
                comment_sheet = CommentSheet.from_tree(fromstring(src))
                for ref, comment in comment_sheet.comments:
                    try:
                        ws[ref].comment = comment
                    except AttributeError:
                        c = ws[ref]
                        if isinstance(c, MergedCell):
                            warnings.warn(comment_warning.format(
                                ws.title, c.coordinate))
                            continue

            # preserve link to VML file if VBA
            if self.wb.vba_archive and ws.legacy_drawing:
                ws.legacy_drawing = rels[ws.legacy_drawing].target

            for t in ws_parser.tables:
                src = self.archive.read(t)
                xml = fromstring(src)
                table = Table.from_tree(xml)
                ws.add_table(table)

            drawings = rels.find(SpreadsheetDrawing._rel_type)
            for rel in drawings:
                charts, images = find_images(self.archive, rel.target)
                for c in charts:
                    ws.add_chart(c, c.anchor)
                for im in images:
                    ws.add_image(im, im.anchor)

            pivot_rel = rels.find(TableDefinition.rel_type)
            for r in pivot_rel:
                pivot_path = r.Target
                src = self.archive.read(pivot_path)
                tree = fromstring(src)
                pivot = TableDefinition.from_tree(tree)
                pivot.cache = self.parser.pivot_caches[pivot.cacheId]
                ws.add_pivot(pivot)

            ws.sheet_state = sheet.state


#重写的 load_workbook ，用于使用自定义的 ExcelReader
def load_workbook(filename, read_only=False, keep_vba=False,
                  data_only=False, keep_links=False):
    print('正在打开工作簿：'+filename)
    reader = ExcelReader(filename, read_only, keep_vba,
                      data_only, keep_links)
    reader.read()

    return reader.wb


