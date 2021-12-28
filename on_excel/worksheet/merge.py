import copy
import openpyxl.worksheet.merge as o_merge
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.styles.borders import Border


from ..cell import Cell,MergedCell


class MergedCellRange(o_merge.MergedCellRange):
    def __init__(self, worksheet, coord):
        from .worksheet import Worksheet
        super().__init__(worksheet, coord)
        self.ws:Worksheet = worksheet
        self.start_cell:Cell=self.start_cell

    def unmerge(self):
        # 这里的行和列的起始值（索引），和Excel的一样，从1开始，并不是从0开始（注意）
        r1, r2, c1, c2 = self.min_row, self.max_row, self.min_col, self.max_col
        self.ws.unmerge_cells(start_row=r1, end_row=r2, start_column=c1, end_column=c2)
        
    def fill(self,value):
        r1, r2, c1, c2 = self.min_row, self.max_row, self.min_col, self.max_col
        # 数据填充
        for r in range(r1, r2+1):  # 遍历行
            if c2 - c1 > 0:  # 多列
                # 遍历列
                for c in range(c1, c2+1):
                    self.ws.cell(r, c).value = value
            else:  # 单列
                self.ws.cell(r, c1).value = value
                self.start_cell
                
    @property
    def shape(self):
        return self.max_col-self.min_col+1,self.max_row-self.min_row+1

    def format(self):
        """
        Each cell of the merged cell is created as MergedCell if it does not
        already exist.

        The MergedCells at the edge of the merged cell gets its borders from
        the upper left cell.

         - The top MergedCells get the top border from the top left cell.
         - The bottom MergedCells get the bottom border from the top left cell.
         - The left MergedCells get the left border from the top left cell.
         - The right MergedCells get the right border from the top left cell.
        """

        names = ['top', 'left', 'right', 'bottom']

        for name in names:
            side = getattr(self.start_cell.border, name)
            if side and side.style is None:
                continue # don't need to do anything if there is no border style
            border = Border(**{name:side})
            for coord in getattr(self, name):
                cell = self.ws._cells.get(coord)
                if cell is None:
                    row, col = coord
                    cell = MergedCell(self.ws, row=row, column=col)
                    self.ws._cells[(cell.row, cell.column)] = cell
                cell.border += border

        protected = self.start_cell.protection is not None
        if protected:
            protection = copy.copy(self.start_cell.protection)
        for coord in self.cells:
            cell = self.ws._cells.get(coord)
            if cell is None:
                row, col = coord
                cell = MergedCell(self.ws, row=row, column=col)
                self.ws._cells[(cell.row, cell.column)] = cell

            if protected:
                cell.protection = protection