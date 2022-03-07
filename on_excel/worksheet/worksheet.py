import typing


import openpyxl.worksheet.worksheet as o_sheet
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter, column_index_from_string
from inspect import isgenerator
from copy import copy
import pandas as pd


from ..cell import Cell, MergedCell
from .merge import MergedCellRange,CellRange


class Worksheet(o_sheet.Worksheet):
    """
    继承了原 Worksheet 的类，为了实现扩展方法
    """

    def __init__(self, parent, title=None) -> None:
        super(Worksheet, self).__init__(parent, title)

    @property
    def visiable(self):
        return self.sheet_state == 'visible'

    @visiable.setter
    def visiable(self, enable: bool):
        if enable:
            self.sheet_state = 'visible'
        else:
            self.sheet_state = 'hidden'

    @property
    def merged_ranges(self):
        m_list:typing.List[MergedCellRange] = list(self.merged_cells)
        return (rg for rg in m_list)

    def merge_cells(self, range_string=None, start_row=None, start_column=None, end_row=None, end_column=None):
        """ Set merge on a cell range.  Range is a cell range (e.g. A1:E1) """
        if range_string is None:
            cr = CellRange(range_string=range_string, min_col=start_column, min_row=start_row,
                      max_col=end_column, max_row=end_row)
            range_string = cr.coord
        mcr = MergedCellRange(self, range_string)
        self.merged_cells.add(mcr)
        self._clean_merge_range(mcr)

    def set_freeze_panes(self, coordinate: str):
        self.freeze_panes = coordinate

    def clear(self, style=True):
        """Clear all content in sheet.

        Args:
            style (bool, optional): is or not clear style . Defaults to True.
        """
        row = 1
        maxrow = self.max_row
        if style:
            self.delete_rows(1, maxrow)
        else:
            self.unmerges(True)
            maxcol = self.max_column
            while row <= maxrow:
                col = 1
                while col <= maxcol:
                    self.cell(row, col).value = ''
                    col += 1
                row += 1

    def match_value_rows(self, value, on_col: str, agg='equal'):
        """Get rows fitted with condition.

        Args:
            value ([type]): [description]
            on_col (str): [description]
            agg (str, optional): 'equal'  'start'  'end'  'in'  'not_in'. Defaults to 'equal'.

        Returns:
            [type]: [description]
        """
        if isinstance(value, str):
            if agg == 'equal':
                return [cell.row for cell in self[on_col] if (cell.value != None) and (cell.value == value)]
            elif agg == 'start':
                return [cell.row for cell in self[on_col] if (cell.value != None) and (cell.value.startswith(value))]
            elif agg == 'in':
                return [cell.row for cell in self[on_col] if (cell.value != None) and (value in cell.value)]
            elif agg == 'end':
                return [cell.row for cell in self[on_col] if (cell.value != None) and (cell.value.endswith(value))]
            elif agg == 'not_in':
                return [cell.row for cell in self[on_col] if (cell.value != None) and (value not in cell.value)]
        elif isinstance(value, list):
            ret = []
            for x in value:
                ret.extend(self.match_value_rows(x, on_col, agg))
            return ret

    def append_df(self, df: pd.DataFrame):
        df.reset_index(drop=False, inplace=True)
        df = df.drop(df.columns.tolist()[0], axis=1)
        self.append(df.columns.values.tolist())
        [self.append(row) for row in df.values.tolist()]

    def unmerge(self, fill):
        # 拆分合并的单元格 并填充内容
        for m_area in self.merged_ranges:
            m_area.unmerge()
            if fill:
                m_area.fill(m_area.start_cell.value)
        return self

    def toDataframe(self, columns_row=0):
        df = pd.DataFrame(self.values)
        # 重设df的columns，并丢掉包含column在内的之前的行
        df.columns = df.iloc[columns_row].tolist()
        df.drop(range(0, columns_row+1), inplace=True, axis=0)  # 丢弃行
        return df

    def merge_col(self, col: str, start_row: int, end_row: int,  align=False):
        """在 col 列合并 从 start_row 到 end_row 之间的单元格

        Args:
            col (str): [description]
            start_row (int): [description]
            end_row (int): [description]
            align (bool, optional): [description]. Defaults to False.

        Returns:
            [type]: [description]
        """
        left_coordinate = "%s%d" % (col, start_row)
        right_coordinate = "%s%d" % (col,  end_row)
        self.merge_cells('%s:%s' % (left_coordinate, right_coordinate))
        if align:
            alignment = Alignment("center", "center", wrap_text=True)
            self[left_coordinate].alignment = alignment
        return self

    def same_value_groups(self, col:str):
        """在 col 列寻找所有值连续相同的单元格，并返回其首尾行的列表

        Args:
            col (str): [description]

        Returns:
            [type]: [description]
        """
        groups = []

        left = right = 0
        cell_arr = self[col]
        max_index = len(cell_arr)-1

        while right <= max_index:
            value_left = cell_arr[left].value
            value_right = cell_arr[right].value
            if value_right == value_left:
                if right == max_index:
                    groups.append((left, right))
                right += 1
            else:
                if right - 1 > left:
                    groups.append((left, right - 1))
                left = right

        return groups

    # 样式属性
    def sys_style(self, tar_sheet):
        rows = self.iter_rows()
        for row in rows:
            for cell in row:
                if isinstance(cell, Cell):
                    cell.copy_style(tar_sheet[cell.coordinate])
                    font = copy(cell.font)
                    font.color="000000"
                    cell.font=font
                    
    # 行列属性
    def set_columns_hide(self, columns, enable: bool):
        if isinstance(columns, str):
            self.column_dimensions[columns].hidden = enable
        elif isinstance(columns, tuple):
            self.column_dimensions.group(
                start=columns[0], end=columns[1], hidden=enable)
        else:
            raise TypeError('columns is a tuple or str')

    def set_column_width(self, column, width):
        letter = None
        if isinstance(column, str):
            letter = column
        elif isinstance(column, int):
            letter = get_column_letter(column)

        if width == 0:
            # print(letter+":"+str(max_len))
            self.column_dimensions[letter].width = 2
        elif width <= 12:
            # print(letter+":"+str(max_len))
            self.column_dimensions[letter].width = width + 2
        elif width <= 50:
            # print(letter+":"+str(max_len))
            self.column_dimensions[letter].width = width + 2
        else:
            # print(letter+":"+str(max_len))
            self.column_dimensions[letter].width = 50
            for cell in self[letter]:
                cell.alignment = Alignment(wrap_text=True)

    def auto_width(self, col_list=None, begin_cal_row=1):
        def get_maxlength(self, col, begin_cal_row):
            """
            获取一个类型为object的Series中的最大占位长度，用于确定导出的xlsx文件的列宽
            col : 表头，也参与比较，解决有时候表头过长的问题
            """
            str_list = [str(cell.value) for cell in self[col] if cell != None]
            str_list = str_list[begin_cal_row-1:]
            len_list = []
            for elem in str_list:
                if elem == "None":
                    continue
                elem_split = list(elem)
                length = 0
                for c in elem_split:
                    if ord(c) <= 256:
                        length += 1
                    else:
                        length += 2
                len_list.append(length)
            if len(len_list) == 0:
                return 0
            return max(len_list)

        loop_range = None
        if col_list == None:
            # 全部列都自动宽度
            max_col = self.max_column
            # 由列数获得对应的列字母
            loop_range = [get_column_letter(i) for i in range(1, max_col+1)]
        elif isinstance(col_list, (tuple, list)):
            loop_range = col_list
        elif isinstance(col_list, str):
            loop_range = (col_list,)
        for letter in loop_range:
            max_len = get_maxlength(self, letter, begin_cal_row)
            self.set_column_width(letter, max_len)

    def get_cell_list(self, slice: str, select: typing.List[int] = None):
        cell_list:typing.List[typing.Union[Cell,MergedCell]] = []
       
        def get_cell(ele):
            if isinstance(ele,(Cell,MergedCell)):
                if select:
                    if (ele.row in select):
                        cell_list.append(ele)
                else:
                    cell_list.append(ele)
            else:
                for e in ele:
                    get_cell(e)

        get_cell(self[slice])
        return cell_list

    # 重写此方法，用于快速获取表格中的单元格
    def __getitem__(self, index)->typing.Union[Cell,typing.Tuple[Cell],typing.Tuple[typing.Tuple[Cell]]]:
        if type(index) == tuple:
            return self.cell(index[0], index[1])
        return super(Worksheet, self).__getitem__(index)

    # 重写此方法，用于获取自定义类型
    def cell(self, row, column, value=None):
        """
        Returns a cell object based on the given coordinates.

        Usage: cell(row=15, column=1, value=5)

        Calling `cell` creates cells in memory when they
        are first accessed.

        :param row: row index of the cell (e.g. 4)
        :type row: int

        :param column: column index of the cell (e.g. 3)
        :type column: int

        :param value: value of the cell (e.g. 5)
        :type value: numeric or time or string or bool or none

        :rtype: openpyxl.cell.cell.Cell
        """

        if row < 1 or column < 1:
            raise ValueError("Row or column values must be at least 1")

        cell = self._get_cell(row, column)
        if value is not None:
            cell.value = value

        return cell

    # 重写此方法，用于获取自定义类型
    def _get_cell(self, row, column):
        """
        Internal method for getting a cell from a worksheet.
        Will create a new cell if one doesn't already exist.
        """
        if not 0 < row < 1048577:
            raise ValueError("Row numbers must be between 1 and 1048576")
        coordinate = (row, column)
        if not coordinate in self._cells:
            cell = Cell(self, row=row, column=column)
            self._add_cell(cell)
        cell = self._cells[coordinate]
        if isinstance(cell, (Cell, MergedCell)):
            return cell
        else:
            raise TypeError("Wrong Cell Type. {}".format(type(cell)))

    # 重写此方法，用于使用自定义的 MergedCell
    def _clean_merge_range(self, mcr):
        """
        Remove all but the top left-cell from a range of merged cells
        and recreate the lost border information.
        Borders are then applied
        """
        cells = mcr.cells
        next(cells) # skip first cell
        for row, col in cells:
            self._cells[row, col] = MergedCell(self, row, col)
        mcr.format()

    # 重写此方法，用于获取自定义类型
    def append(self, iterable):
        """Appends a group of values at the bottom of the current sheet.

        * If it's a list: all values are added in order, starting from the first column
        * If it's a dict: values are assigned to the columns indicated by the keys (numbers or letters)

        :param iterable: list, range or generator, or dict containing values to append
        :type iterable: list|tuple|range|generator or dict

        Usage:

        * append(['This is A1', 'This is B1', 'This is C1'])
        * **or** append({'A' : 'This is A1', 'C' : 'This is C1'})
        * **or** append({1 : 'This is A1', 3 : 'This is C1'})

        :raise: TypeError when iterable is neither a list/tuple nor a dict

        """
        row_idx = self._current_row + 1

        if (isinstance(iterable, (list, tuple, range))
                or isgenerator(iterable)):
            for col_idx, content in enumerate(iterable, 1):
                if isinstance(content, Cell):
                    # compatible with write-only mode
                    cell = content
                    if cell.parent and cell.parent != self:
                        raise ValueError(
                            "Cells cannot be copied from other worksheets")
                    cell.parent = self
                    cell.column = col_idx
                    cell.row = row_idx
                else:
                    cell = Cell(self, row=row_idx,
                                column=col_idx, value=content)
                self._cells[(row_idx, col_idx)] = cell

        elif isinstance(iterable, dict):
            for col_idx, content in iterable.items():
                if isinstance(col_idx, str):
                    col_idx = column_index_from_string(col_idx)
                cell = Cell(self, row=row_idx, column=col_idx, value=content)
                self._cells[(row_idx, col_idx)] = cell

        else:
            self._invalid_row(iterable)

        self._current_row = row_idx
